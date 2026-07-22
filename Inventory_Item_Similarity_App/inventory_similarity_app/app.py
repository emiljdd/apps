from __future__ import annotations

import io
import math
import re
from collections import defaultdict
from typing import Callable

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from rapidfuzz import fuzz, process

st.set_page_config(page_title="Inventory Item Similarity App", page_icon="📦", layout="wide")

SCORERS: dict[str, Callable] = {
    "Standard ratio": fuzz.ratio,
    "Partial ratio": fuzz.partial_ratio,
    "Token sort ratio": fuzz.token_sort_ratio,
    "Token set ratio": fuzz.token_set_ratio,
    "Weighted ratio": fuzz.WRatio,
}


def clean_header(value: object) -> str:
    return re.sub(r"\s+", " ", str(value).strip())


def normalize_item(
    value: object,
    uppercase: bool,
    remove_spaces: bool,
    remove_hyphens: bool,
    remove_punctuation: bool,
    remove_leading_zeros: bool,
) -> str:
    if pd.isna(value):
        return ""
    text = str(value).strip()
    if uppercase:
        text = text.upper()
    if remove_spaces:
        text = re.sub(r"\s+", "", text)
    if remove_hyphens:
        text = text.replace("-", "")
    if remove_punctuation:
        text = re.sub(r"[^A-Za-z0-9]", "", text)
    if remove_leading_zeros:
        text = re.sub(r"^0+(?=\d)", "", text)
    return text


@st.cache_data(show_spinner=False)
def get_excel_sheets(file_bytes: bytes) -> list[str]:
    return pd.ExcelFile(io.BytesIO(file_bytes)).sheet_names


@st.cache_data(show_spinner=False)
def read_source(file_bytes: bytes, file_name: str, sheet_name: str | None, header_row: int) -> pd.DataFrame:
    suffix = file_name.lower().rsplit(".", 1)[-1]
    buffer = io.BytesIO(file_bytes)
    if suffix in {"xlsx", "xlsm", "xls"}:
        return pd.read_excel(buffer, sheet_name=sheet_name or 0, header=header_row - 1)
    if suffix == "csv":
        return pd.read_csv(buffer, header=header_row - 1)
    raise ValueError("Supported formats are XLSX, XLSM, XLS, and CSV.")


def aggregate_inventory(
    df: pd.DataFrame,
    item_column: str,
    qty_column: str,
    aggregation: str,
    blank_qty_as_zero: bool,
    quantity_filter: str,
    group_case_sensitive: bool,
    trim_item_numbers: bool,
) -> tuple[pd.DataFrame, int]:
    work = df.copy()
    work.columns = [clean_header(c) for c in work.columns]
    work[item_column] = work[item_column].astype("string")
    if trim_item_numbers:
        work[item_column] = work[item_column].str.strip()

    work = work[work[item_column].notna() & work[item_column].ne("")].copy()
    work[qty_column] = pd.to_numeric(work[qty_column], errors="coerce")
    if blank_qty_as_zero:
        work[qty_column] = work[qty_column].fillna(0)
    else:
        work = work[work[qty_column].notna()].copy()

    source_rows = len(work)
    work["_GroupKey"] = work[item_column] if group_case_sensitive else work[item_column].str.upper()

    agg_map = {
        "Sum": "sum",
        "Maximum": "max",
        "Minimum": "min",
        "Average": "mean",
        "First": "first",
    }
    grouped = (
        work.groupby("_GroupKey", as_index=False)
        .agg({item_column: "first", qty_column: agg_map[aggregation]})
        .rename(columns={item_column: "Item Number", qty_column: f"{aggregation} QTY On Hand"})
    )
    qty_out = f"{aggregation} QTY On Hand"

    if quantity_filter == "Exclude zero totals":
        grouped = grouped[grouped[qty_out].ne(0)]
    elif quantity_filter == "Keep positive totals only":
        grouped = grouped[grouped[qty_out].gt(0)]
    elif quantity_filter == "Keep negative totals only":
        grouped = grouped[grouped[qty_out].lt(0)]

    return grouped.drop(columns="_GroupKey").reset_index(drop=True), source_rows


def allowed_length_range(query_length: int, threshold: float) -> tuple[int, int]:
    t = threshold / 100.0
    min_length = max(1, math.ceil((t / (2 - t)) * query_length))
    max_length = max(query_length, math.floor(((2 - t) / t) * query_length))
    return min_length, max_length


def find_closest_matches(
    summary: pd.DataFrame,
    threshold: float,
    scorer_name: str,
    uppercase: bool,
    remove_spaces: bool,
    remove_hyphens: bool,
    remove_punctuation: bool,
    remove_leading_zeros: bool,
    minimum_length: int,
) -> pd.DataFrame:
    result = summary.copy()
    result["Normalized Item"] = result["Item Number"].map(
        lambda x: normalize_item(
            x, uppercase, remove_spaces, remove_hyphens, remove_punctuation, remove_leading_zeros
        )
    )
    result = result[result["Normalized Item"].str.len().ge(minimum_length)].reset_index(drop=True)

    normalized = result["Normalized Item"].tolist()
    display = result["Item Number"].astype(str).tolist()
    scorer = SCORERS[scorer_name]
    by_length: dict[int, list[int]] = defaultdict(list)
    for i, value in enumerate(normalized):
        by_length[len(value)].append(i)

    closest, scores, flags = [], [], []
    progress = st.progress(0, text="Comparing item numbers...")
    total = max(len(normalized), 1)

    for i, query in enumerate(normalized):
        if scorer_name == "Standard ratio":
            low, high = allowed_length_range(len(query), threshold)
            candidates = [idx for length in range(low, high + 1) for idx in by_length.get(length, []) if idx != i]
        else:
            candidates = [idx for idx in range(len(normalized)) if idx != i]

        choices = [normalized[idx] for idx in candidates]
        best = process.extractOne(query, choices, scorer=scorer, score_cutoff=threshold) if choices else None
        if best:
            _, score, pos = best
            closest.append(display[candidates[pos]])
            scores.append(round(float(score), 2))
            flags.append("Review")
        else:
            closest.append("")
            scores.append(None)
            flags.append("")

        if i % max(1, total // 100) == 0 or i == total - 1:
            progress.progress((i + 1) / total, text=f"Compared {i + 1:,} of {len(normalized):,} items")
    progress.empty()

    result["Closest Similar Item"] = closest
    result["Similarity %"] = scores
    result["Review Flag"] = flags
    return result.drop(columns="Normalized Item")


def build_excel(result: pd.DataFrame, settings: pd.DataFrame, highlight_color: str) -> bytes:
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        result.to_excel(writer, sheet_name="Item Summary", index=False)
        settings.to_excel(writer, sheet_name="Settings", index=False)

    buffer.seek(0)
    wb = load_workbook(buffer)
    ws = wb["Item Summary"]
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    widths = [28, 20, 28, 16, 14]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + idx)].width = width
    if len(result):
        ref = f"A1:E{len(result)+1}"
        tab = Table(displayName="ItemSummaryTable", ref=ref)
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(tab)
        ws.conditional_formatting.add(
            f"A2:E{len(result)+1}",
            FormulaRule(formula=['$E2="Review"'], fill=PatternFill("solid", fgColor=highlight_color)),
        )
    for cell in ws["B"][1:]:
        cell.number_format = "#,##0.00"
    for cell in ws["D"][1:]:
        cell.number_format = "0.00"

    settings_ws = wb["Settings"]
    for cell in settings_ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
    settings_ws.column_dimensions["A"].width = 32
    settings_ws.column_dimensions["B"].width = 55

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


st.title("Inventory Item Similarity App")
st.caption("Aggregate item quantities, remove zero totals, identify near-duplicate item numbers, and export a formatted workbook.")

uploaded = st.file_uploader("Upload an Excel or CSV file", type=["xlsx", "xlsm", "xls", "csv"])

with st.sidebar:
    st.header("Processing parameters")
    header_row = st.number_input("Header row number", min_value=1, value=1, step=1)
    aggregation = st.selectbox("Quantity aggregation", ["Sum", "Maximum", "Minimum", "Average", "First"])
    quantity_filter = st.selectbox(
        "Quantity filter",
        ["Exclude zero totals", "Keep positive totals only", "Keep negative totals only", "Keep all totals"],
    )
    blank_qty_as_zero = st.checkbox("Treat blank/non-numeric quantity as zero", value=True)
    trim_items = st.checkbox("Trim leading/trailing spaces", value=True)
    group_case_sensitive = st.checkbox("Case-sensitive item grouping", value=False)

    st.header("Similarity parameters")
    threshold = st.slider("Similarity threshold (%)", 50.0, 100.0, 95.0, 0.5)
    scorer_name = st.selectbox("Similarity method", list(SCORERS))
    minimum_length = st.number_input("Minimum normalized item length", min_value=1, value=3, step=1)
    uppercase = st.checkbox("Ignore capitalization", value=True)
    remove_spaces = st.checkbox("Ignore spaces", value=True)
    remove_hyphens = st.checkbox("Ignore hyphens", value=True)
    remove_punctuation = st.checkbox("Ignore all punctuation", value=True)
    remove_leading_zeros = st.checkbox("Ignore leading zeros", value=False)

    st.header("Export parameters")
    output_name = st.text_input("Output file name", value="Item_Summary_Similarity.xlsx")
    highlight_color = st.color_picker("Review highlight color", value="#FFF2CC").replace("#", "")
    preview_rows = st.number_input("Preview rows", min_value=10, max_value=500, value=100, step=10)

if uploaded:
    file_bytes = uploaded.getvalue()
    is_excel = uploaded.name.lower().endswith((".xlsx", ".xlsm", ".xls"))
    sheet_name = None
    if is_excel:
        sheets = get_excel_sheets(file_bytes)
        sheet_name = st.selectbox("Worksheet", sheets)

    try:
        source = read_source(file_bytes, uploaded.name, sheet_name, int(header_row))
        source.columns = [clean_header(c) for c in source.columns]
        if source.empty:
            st.error("The selected source contains no data rows.")
            st.stop()

        c1, c2 = st.columns(2)
        with c1:
            item_column = st.selectbox("Item number column", list(source.columns), index=0)
        with c2:
            default_qty_index = 1 if len(source.columns) > 1 else 0
            qty_column = st.selectbox("Quantity column", list(source.columns), index=default_qty_index)

        with st.expander("Source preview", expanded=False):
            st.dataframe(source.head(int(preview_rows)), use_container_width=True)

        if st.button("Process inventory", type="primary", use_container_width=True):
            with st.spinner("Aggregating quantities..."):
                summary, valid_source_rows = aggregate_inventory(
                    source, item_column, qty_column, aggregation, blank_qty_as_zero,
                    quantity_filter, group_case_sensitive, trim_items,
                )
            if summary.empty:
                st.warning("No item numbers remain after applying the selected filters.")
                st.stop()

            result = find_closest_matches(
                summary, threshold, scorer_name, uppercase, remove_spaces,
                remove_hyphens, remove_punctuation, remove_leading_zeros,
                int(minimum_length),
            )
            flagged = int(result["Review Flag"].eq("Review").sum())

            m1, m2, m3, m4 = st.columns(4)
            m1.metric("Source rows used", f"{valid_source_rows:,}")
            m2.metric("Unique items", f"{len(result):,}")
            m3.metric("Items flagged", f"{flagged:,}")
            m4.metric("Flag rate", f"{flagged / len(result):.1%}" if len(result) else "0.0%")

            st.dataframe(result.head(int(preview_rows)), use_container_width=True)

            settings = pd.DataFrame({
                "Parameter": [
                    "Source file", "Worksheet", "Header row", "Item column", "Quantity column",
                    "Aggregation", "Quantity filter", "Blank quantity as zero", "Case-sensitive grouping",
                    "Trim item numbers", "Similarity threshold", "Similarity method", "Minimum item length",
                    "Ignore capitalization", "Ignore spaces", "Ignore hyphens", "Ignore punctuation",
                    "Ignore leading zeros",
                ],
                "Value": [
                    uploaded.name, sheet_name or "N/A", header_row, item_column, qty_column,
                    aggregation, quantity_filter, blank_qty_as_zero, group_case_sensitive,
                    trim_items, threshold, scorer_name, minimum_length, uppercase, remove_spaces,
                    remove_hyphens, remove_punctuation, remove_leading_zeros,
                ],
            })
            excel_bytes = build_excel(result, settings, highlight_color)
            csv_bytes = result.to_csv(index=False).encode("utf-8-sig")

            d1, d2 = st.columns(2)
            d1.download_button(
                "Download formatted Excel workbook", excel_bytes,
                file_name=output_name if output_name.lower().endswith(".xlsx") else output_name + ".xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
            d2.download_button(
                "Download CSV", csv_bytes,
                file_name=re.sub(r"\.xlsx$", ".csv", output_name, flags=re.I),
                mime="text/csv", use_container_width=True,
            )
    except Exception as exc:
        st.exception(exc)
else:
    st.info("Upload a source file to begin. All parameters can be changed from the sidebar.")
